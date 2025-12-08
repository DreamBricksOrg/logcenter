from __future__ import annotations

import io
import csv
import json
import openpyxl
from openpyxl.utils import get_column_letter
from typing import List, Optional, Dict, Any
from bson import ObjectId

from db.utils import get_db
from util.helpers import utcnow_iso
from services.stream_service import manager
from util.queries import (
    build_visibility_match,
    merge_filters,
    convert_timestamp_filters,
    cap_limit,
)

DEFAULT_LIMIT = 1000
MAX_LIMIT = 10000


def _serialize_ids(doc: Dict[str, Any]) -> Dict[str, Any]:
    """
    Normaliza ObjectId -> str em _id e project_id para resposta JSON.
    Também garante presença de 'status' e 'uploadedAt' na saída.
    """
    out = dict(doc)
    for k in ["_id", "project_id"]:
        if k in out and isinstance(out[k], ObjectId):
            out[k] = str(out[k])
    if "status" not in out:
        out["status"] = str(out.get("level", "INFO")).upper()
    if "uploadedAt" not in out:
        out["uploadedAt"] = utcnow_iso()
    return out


async def _ensure_project_active(db, project_id: str) -> ObjectId:
    """
    Valida project_id (string OID) e garante que o projeto existe e está ATIVO.
    Se estiver inativo (ou não existir), retorna erro de forma a não vazar existência.
    """
    try:
        oid = ObjectId(project_id)
    except Exception:
        raise ValueError("Invalid project_id (must be a valid ObjectId string)")

    proj = await db["projects"].find_one({"_id": oid, "status": "active"}, {"_id": 1})
    if not proj:
        raise ValueError("Project not found for given project_id")  # oculta inativos
    return oid


async def create_log(
    project_id: str,
    status: str,
    level: str,
    message: str,
    timestamp: Optional[str] = None,
    tags: Optional[List[str]] = None,
    data: Optional[Dict[str, Any]] = None,
    request_id: Optional[str] = None,
) -> str:
    """
    Cria e insere log no db + emite broadcast por canal do projeto.
    Bloqueia logs de projetos inativos (tratados como não encontrados).
    """
    db = await get_db()

    proj_oid = await _ensure_project_active(db, project_id)
    uploaded_at = utcnow_iso()
    event_ts = timestamp or utcnow_iso()

    log_doc: Dict[str, Any] = {
        "uploadedAt": uploaded_at,
        "timestamp": event_ts,
        "status": status,
        "level": level,
        "message": message,
        "tags": tags or [],
        "data": data or {},
        "request_id": request_id or None,
        "project_id": proj_oid,
    }

    res = await db["logs"].insert_one(log_doc)

    # Broadcast em tempo real por canal do project_id (string)
    await manager.broadcast(
        str(proj_oid),
        json.dumps(
            {**_serialize_ids(log_doc), "_id": str(res.inserted_id)},
            ensure_ascii=False,
            separators=(",", ":"),
        ),
    )

    return str(res.inserted_id)


async def list_logs(
    filters: Optional[Dict[str, Any]] = None,
    project_id: Optional[str] = None,
    visibility: Optional[Dict[str, Any]] = None,
    limit: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """
    Lista logs filtrados mais recentes.
    Aplica visibilidade (espera receber filtros por project_id) e filtro opcional por project_id.
    - `limit` opcional com teto (DEFAULT_LIMIT=1000, MAX_LIMIT=10000).
    - Sempre restringe a projetos ATIVOS via visibilidade.
    """
    db = await get_db()

    match = await build_visibility_match(visibility, project_id)
    match = merge_filters(match, filters or {})
    match = convert_timestamp_filters(match)

    n = cap_limit(limit, DEFAULT_LIMIT, MAX_LIMIT)
    cursor = db["logs"].find(match).sort("timestamp", -1).limit(n)
    docs = await cursor.to_list(length=n)
    return [_serialize_ids(d) for d in docs]


async def latest_timestamp(project_id: Optional[str] = None) -> Optional[str]:
    """
    Retorna o timestamp mais recente de um projeto (ou geral),
    SEMPRE considerando apenas projetos ATIVOS.
    """
    db = await get_db()

    match = await build_visibility_match(visibility=None, project_id=project_id)

    doc = await db["logs"].find_one(
        match,
        sort=[("timestamp", -1)],
        projection={"timestamp": 1},
    )
    return doc["timestamp"] if doc else None


async def level_counts(project_id: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Conta logs agrupados por 'level' (normalizado para UPPER), SOMENTE de projetos ATIVOS.
    """
    db = await get_db()

    match = await build_visibility_match(visibility=None, project_id=project_id)

    pipeline: List[Dict[str, Any]] = [
        {"$match": match},
        {"$addFields": {"_norm_level": {"$toUpper": "$level"}}},
        {"$group": {"_id": "$_norm_level", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}},
        {"$project": {"level": "$_id", "_id": 0, "count": 1}},
    ]

    cursor = db["logs"].aggregate(pipeline)
    return [doc async for doc in cursor]


async def generate_logs_csv(
    filters: Optional[Dict[str, Any]] = None,
    visibility: Optional[Dict[str, Any]] = None,
    project_id: Optional[str] = None,
    limit: Optional[int] = None,
) -> io.BytesIO:
    """
    Exporta logs filtrados como CSV (UTF-8). Sempre somente projetos ATIVOS.
    - `project_id` opcional (string ObjectId) para filtrar explicitamente.
    - `limit` opcional (sem teto). Se None, exporta tudo.
    """
    db = await get_db()

    match = await build_visibility_match(visibility, project_id)
    match = merge_filters(match, filters or {})
    match = convert_timestamp_filters(match)

    projection = {
        "_id": 1, "uploadedAt": 1, "timestamp": 1, "status": 1, "level": 1,
        "message": 1, "tags": 1, "data": 1, "request_id": 1, "project_id": 1,
    }

    cursor = db["logs"].find(match, projection).sort("timestamp", -1)
    if isinstance(limit, int) and limit > 0:
        cursor = cursor.limit(int(limit))

    output = io.StringIO()
    fieldnames = ["_id","uploadedAt","timestamp","status","level","message","tags","data","request_id","project_id"]

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    async for d in cursor:
        d = _serialize_ids(d)
        tags_val = d.get("tags") or []
        tags_str = ";".join(map(str, tags_val)) if isinstance(tags_val, list) else str(tags_val)

        data_val = d.get("data") or {}
        try:
            data_json = json.dumps(data_val, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            data_json = "{}"

        writer.writerow(
            {
                "_id": d.get("_id", ""),
                "uploadedAt": d.get("uploadedAt", ""),
                "timestamp": d.get("timestamp", ""),
                "status": d.get("status", ""),
                "level": d.get("level", ""),
                "message": d.get("message", ""),
                "tags": tags_str,
                "data": data_json,
                "request_id": d.get("request_id", "") or "",
                "project_id": d.get("project_id", ""),
            }
        )

    output.seek(0)
    return io.BytesIO(output.getvalue().encode("utf-8"))


async def generate_logs_excel(
    filters: Optional[Dict[str, Any]] = None,
    visibility: Optional[Dict[str, Any]] = None,
    project_id: Optional[str] = None,
    limit: Optional[int] = None,
) -> io.BytesIO:
    """
    Exporta logs em Excel (.xlsx). Sempre somente projetos ATIVOS.
    - `project_id` opcional (string ObjectId).
    - `limit` opcional (sem teto). Se None, exporta tudo.
    """
    db = await get_db()

    match = await build_visibility_match(visibility, project_id)
    match = merge_filters(match, filters or {})
    match = convert_timestamp_filters(match)

    cursor = db["logs"].find(match).sort("timestamp", -1)
    if isinstance(limit, int) and limit > 0:
        cursor = cursor.limit(int(limit))

    docs = [_serialize_ids(d) async for d in cursor]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs"

    headers = ["_id","uploadedAt","timestamp","status","level","message","tags","data","request_id","project_id"]

    ws.append(headers)

    for d in docs:
        tags_val = d.get("tags") or []
        tags_str = ";".join(map(str, tags_val)) if isinstance(tags_val, list) else str(tags_val)
        data_json = json.dumps(d.get("data") or {}, ensure_ascii=False, separators=(",", ":"))
        ws.append(
            [
                d.get("_id", ""),
                d.get("uploadedAt", ""),
                d.get("timestamp", ""),
                d.get("status", ""),
                d.get("level", ""),
                d.get("message", ""),
                tags_str,
                data_json,
                d.get("request_id", ""),
                d.get("project_id", ""),
            ]
        )

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
